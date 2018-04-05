import itchat
from itchat.content import *
import threading
import queue
import time
import openpyxl


class Worker(threading.Thread):
    def __init__(self, message_queue, menu, lock):
        threading.Thread.__init__(self)

        self.message_queue = message_queue
        # 菜单
        self.menu = menu

        # 单次食品列表
        self.foods = []
        # 单次订餐价格
        self.price = 0
        # 单次订餐地址
        self.address = None

        # 锁
        self.lock = lock
        # 具体的一条信息
        self.message = None
        self.setDaemon(True)
        self.start()

    def run(self):
        while True:
            try:
                # 从队列中取出一条消息
                self.message = self.message_queue.get()
            except self.message_queue.Empty:
                continue

            # 做匹配
            if self.message.text == '菜单':
                self.message.user.send(str(self.menu))

            elif self.message.text == '帮助':
                self.message.user.send('下单格式如下：\n黄焖鸡米饭 地址 小胡同路22号')

            elif self.match():
                # 匹配成功，存入excel中
                # print('当前用户', self.message.user)
                if self.message.user['UserName'] != 'filehelper':
                #if True:
                    with self.lock:
                        self.save()
                    self.message.user.send('下单成功')

                print('下单 UserName：', self.message.ToUserName)
                print('价格 Price：', self.price)
                print('foods：', self.foods)
                print('地址：', self.address)
            else:
                self.message.user.send("输入'菜单'，获取菜单\n" + "输入'帮助'，获取帮助\n")
                # logging.DEBUG('消息来自',self.message.FromUserName)
            self.message_queue.task_done()

    def match(self):
        for item in self.menu:
            if item in self.message.text and '地址' in self.message.text:
                print('匹配成功')
                with self.lock:
                    self.foods.append(item)
                    self.price += menu[item]

        if self.price != 0:
            with self.lock:
                self.address = self.message.text.split('地址')[-1]
            return (self.foods, self.price, self.address)
        return False

    def save(self):
        # 新建工作薄
        workbook = openpyxl.load_workbook(filename)
        # 新建工作表
        sheet = workbook.active
        # 写入数据
        sheet.append([self.message.user['NickName'], self.foods, self.price, self.address, time.ctime(), self.message.text])
        # 测试filehelp
        #sheet.append([self.message.user['UserName'], str(self.foods), self.price, self.address, time.ctime(), self.message.text])
        # 保存文件
        workbook.save(filename)


class ThreadPool:

    def __init__(self, num, menu, lock):
        self.num = num
        self.queue = queue.Queue()
        self.threads = []
        self.menu = menu
        self.lock = lock
        self.create_thread()

    def create_thread(self):
        for i in range(self.num):
            self.threads.append(Worker(self.queue, self.menu, self.lock))

    def put_job(self, message):
        self.queue.put(message)

    def wait(self):
        self.queue.join()


# 注册消息
@itchat.msg_register([TEXT, MAP, CARD, NOTE, SHARING])
def get_message(msg):
    # logging.DEBUG('消息发往：', msg.ToUserName)
    # logging.DEBUG('消息用户：', msg.user)
    # logging.DEBUG('收到消息')
    #logging.DEBUG(str(itchat.web_init()['User']['NickName']))
    pool.put_job(msg)


def get_menu():
    print('读取菜单')
    # 读取文件menu.txt菜单
    #try:
    #    with open('menu.txt', 'r') as f:
    #        menu = f.read().split('\n')[:-1]
    #    print('读取成功')
    #    print(str(menu))
    #except IOError:
    #    print('读取失败')
    #    return
    menu = dict()
    try:
        wb = openpyxl.load_workbook('menu.xlsx', read_only=True)
        ws = wb.active
        for row in range(ws.max_row):
            menu[ws.cell(row + 1, 1).value] = ws.cell(row + 1, 2).value
        print(str(menu))

    except IOError:
        print('读取失败')

    return menu


def init_excel(filename):
    # 新建工作薄
    workbook = openpyxl.Workbook()
    # 新建工作表
    ws = workbook.active
    # 写入数据
    ws['A1'] = '顾客'
    ws['B1'] = '菜品'
    ws['C1'] = '价格'
    ws['D1'] = '地址'
    ws['E1'] = '时间'
    ws['F1'] = '消息'
    # 保存文件
    workbook.save(filename)



if __name__ == '__main__':
    # 线程数
    num = 5
    # 菜单
    menu = get_menu()

    lock = threading.Lock()

    if not menu:
        exit()

    print('线程初始化')
    pool = ThreadPool(num, menu, lock)

    print('初始化工作薄')
    filename = time.strftime("%Y-%m-%d %H:%M", time.localtime())+ '.xlsx'
    init_excel(filename)

    print('登录中……')
    itchat.auto_login(enableCmdQR=2, hotReload=True)
    print('登录成功')
    itchat.run(True)

    pool.wait()




